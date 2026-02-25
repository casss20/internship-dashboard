import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime

# Load workbook
wb = openpyxl.load_workbook("/root/.openclaw/workspace/Anthony_Internship_Tracker_2026.xlsx")
ws = wb["Applications"]

# Get current row count
start_row = ws.max_row + 1

# New opportunities found
new_opportunities = [
    ["", "Intel", "Graduate Software Engineering Intern", "https://jobs.intel.com", "Direct", "Summer 2026", "Oregon", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Top tech company - 🔥 HIGH PRIORITY", "88", "High", "", ""],
    ["", "Walmart", "Software Engineering II Intern", "https://careers.walmart.com", "Direct", "Summer 2026", "Sunnyvale, CA", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Fortune 1 company - 🔥 HIGH PRIORITY", "87", "High", "", ""],
    ["", "T-Mobile", "AI Engineering Intern", "https://careers.t-mobile.com", "Direct", "Summer 2026", "Washington", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Telecom + AI - 🔥 HIGH PRIORITY", "86", "High", "", ""],
    ["", "Zoox", "AI Applications Development Intern", "https://zoox.com/careers", "Direct", "Summer 2026", "Foster City, CA", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Autonomous vehicles (Amazon) - 🔥 HIGH PRIORITY", "89", "High", "", ""],
    ["", "Toyota Research Institute", "Human-Centered AI Intern", "https://tri.global/careers", "Direct", "Summer 2026", "Los Altos, CA", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Research + AI - 🔥 HIGH PRIORITY", "88", "High", "", ""],
    ["", "Visa", "Cybersecurity Applied Scientist Intern", "https://usa.visa.com/careers", "Direct", "Summer 2026", "Ashburn, VA", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Cyber + AI - 🔥 HIGH PRIORITY", "90", "High", "", ""],
    ["", "Samsung SDS", "AI Platform Lab Research Intern", "https://www.samsungsds.com/careers", "Direct", "Summer 2026", "Mountain View, CA", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Samsung AI research - 🔥 HIGH PRIORITY", "87", "High", "", ""],
    ["", "Northrop Grumman", "AI Engineer Intern", "https://www.northropgrumman.com/careers", "Direct", "Summer 2026", "Remote", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Defense + AI - 🔥 HIGH PRIORITY", "85", "High", "", ""],
    ["", "insitro", "Machine Learning Intern", "https://www.insitro.com/careers", "Direct", "Summer 2026", "South San Francisco, CA", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Biotech + AI - 🔥 HIGH PRIORITY", "88", "High", "", ""],
    ["", "Sigma Computing", "Software Engineering Intern", "https://www.sigmacomputing.com/careers", "Direct", "Summer 2026", "San Francisco, CA", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "AI/Analytics - 🔥 HIGH PRIORITY", "85", "High", "", ""],
    ["", "Applied Materials", "Software Engineer Intern", "https://www.appliedmaterials.com/careers", "Direct", "Summer 2026", "Santa Clara, CA", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Semiconductor leader - 🔥 HIGH PRIORITY", "84", "High", "", ""],
    ["", "Okta", "Data Analyst Intern", "https://www.okta.com/careers", "Direct", "Summer 2026", "San Francisco, CA", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Identity/Security - 🔥 HIGH PRIORITY", "83", "High", "", ""],
    ["", "Scopely", "Software Engineering Intern", "https://www.scopely.com/careers", "Direct", "Summer 2026", "Bellevue, WA / San Francisco, CA", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Gaming (Pokémon GO) - 🔥 HIGH PRIORITY", "84", "High", "", ""],
    ["", "Raytheon", "Software Engineering Intern", "https://careers.raytheon.com", "Direct", "Summer 2026", "Various", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Defense contractor", "82", "High", "", ""],
    ["", "Motorola Solutions", "Software Test Engineer Intern", "https://www.motorolasolutions.com/careers", "Direct", "Summer 2026", "California", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Communications tech", "80", "Medium", "", ""],
    ["", "GE HealthCare", "Software Engineering Intern", "https://www.gehealthcare.com/careers", "Direct", "Summer 2026", "Ohio", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Healthcare tech", "79", "Medium", "", ""],
    ["", "Micron Technology", "SSD Firmware Intern", "https://www.micron.com/careers", "Direct", "Summer 2026", "San Jose, CA", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Memory/Storage", "81", "High", "", ""],
    ["", "Roku", "Software Engineer Intern", "https://www.roku.com/careers", "Direct", "Summer 2026", "UK / Remote", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Streaming platform", "78", "Medium", "", ""],
    ["", "Credit Acceptance", "Software Engineer Intern", "https://www.creditacceptance.com/careers", "Direct", "Summer 2026", "Remote", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Fintech", "77", "Medium", "", ""],
    ["", "Roadie", "Software Engineer Intern", "https://www.roadie.com/careers", "Direct", "Summer 2026", "Remote", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Delivery platform", "76", "Medium", "", ""],
    ["", "PTC", "Software Engineering Intern", "https://www.ptc.com/careers", "Direct", "Summer 2026", "Boston, MA", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Industrial software", "75", "Medium", "", ""],
    ["", "KLA", "Research Scientist Intern", "https://www.kla.com/careers", "Direct", "Summer 2026", "Ann Arbor, MI", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Semiconductor research", "80", "High", "", ""],
    ["", "Johnson & Johnson", "Clinical Data Analytics & AI Intern", "https://www.jnj.com/careers", "Direct", "Summer 2026", "Cincinnati, OH", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "Healthcare AI", "79", "Medium", "", ""],
    ["", "Bloom (YC)", "AI/Full-Stack Intern", "https://www.ycombinator.com/companies/bloom-4", "Y Combinator", "Summer 2026", "Zurich", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "$6-10K/month", "78", "Medium", "", ""],
    ["", "Cohere", "Software Engineering", "https://cohere.com/careers", "Direct", "Summer 2026", "Remote", "Not Started", "Passed", "", "No", "No", "", "", "No", "", "No", "", "AI company", "80", "Medium", "", ""],
]

# Color fills
fill_yellow = PatternFill(start_color="f5a623", end_color="f5a623", fill_type="solid")

# Add Application IDs and data
app_id = 16  # Starting after existing 15
for i, opp in enumerate(new_opportunities, start=start_row):
    opp[0] = f"APP-{app_id:03d}"  # Application ID
    app_id += 1
    for col_num, value in enumerate(opp, 1):
        cell = ws.cell(row=i, column=col_num)
        cell.value = value
        # Color status column
        if col_num == 9:  # Status column
            cell.fill = fill_yellow
            cell.font = Font(color="000000")

# Update Summary sheet
if "Summary" in wb.sheetnames:
    ws_summary = wb["Summary"]
    ws_summary['B4'] = ws.max_row - 1  # Total applications
    ws_summary['B12'] = '=COUNTIF(Applications!W:W,"High")'  # High priority
    ws_summary['B13'] = '=COUNTIF(Applications!W:W,"Medium")'  # Medium priority
    ws_summary['B15'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Save
wb.save("/root/.openclaw/workspace/Anthony_Internship_Tracker_2026.xlsx")

print(f"✅ Added {len(new_opportunities)} new opportunities to tracker")
print(f"📊 Total applications now: {ws.max_row - 1}")
print(f"🔥 High priority: 13 new")
