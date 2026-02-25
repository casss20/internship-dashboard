import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import json

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Applications"

# Define columns
columns = [
    "Date Applied", "Company", "Role Title", "Job URL", "Application Portal",
    "Term", "Location", "Status", "Resume Version Used", "Cover Letter",
    "Cold Email Sent", "Recruiter Name", "Recruiter Email", "Cold Email Date",
    "Follow-up Sent", "Follow-up Date", "Response Received", "Response Type",
    "Notes", "Match Score (%)", "Priority"
]

# Header row
for col_num, column_title in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = column_title
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1f2530", end_color="1f2530", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Seed data from your tracker
seed_data = [
    ["", "SupplyHouse", "AI Enablement", "https://www.supplyhouse.com/careers", "Direct", "Summer 2026", "Remote", "Ready to Submit", "AI_Enhanced", "Yes", "No", "", "", "", "No", "", "No", "", "Cover letter ready", "85", "High"],
    ["", "Mercury", "Data Science", "https://mercury.com/careers", "Direct", "Summer 2026", "Remote", "Ready to Submit", "Data_Science", "Yes", "No", "", "", "", "No", "", "No", "", "Cover letter ready", "82", "High"],
    ["", "Edged", "2026 Summer Intern", "", "Direct", "Summer 2026", "Remote", "Ready to Submit", "General", "Yes", "No", "", "", "", "No", "", "No", "", "Cover letter ready", "80", "High"],
    ["", "Scale AI", "ML Research Intern", "https://scale.com/careers", "Direct", "Summer 2026", "SF/Remote", "Not Started", "", "Yes", "No", "Jonathan Taylor", "jonathan@scale.com", "", "No", "", "No", "", "Strong AI/ML fit", "88", "High"],
    ["", "DoorDash", "Machine Learning Intern", "https://careers.doordash.com", "Direct", "Summer 2026", "SF/NY/Seattle", "Not Started", "", "Yes", "No", "", "", "", "No", "", "No", "", "ML pipeline experience match", "85", "High"],
    ["", "GuidePoint", "GPSU Cybersecurity", "https://www.guidepointsecurity.com/careers", "Direct", "Summer 2026", "Remote", "Ready to Submit", "Cybersecurity", "Yes", "No", "Taylor Pierotti", "taylor@guidepointsecurity.com", "", "No", "", "No", "", "Cover letter ready", "90", "High"],
    ["", "Crystal Dynamics", "Cyber Security", "https://crystaldynamics.com/careers", "Direct", "Summer 2026", "Hybrid", "Ready to Submit", "Cybersecurity", "Yes", "No", "", "", "", "No", "", "No", "", "Cover letter ready", "87", "High"],
    ["", "CrowdStrike", "IT Service Desk Launch Intern", "https://www.crowdstrike.com/careers", "Direct", "Summer 2026", "Austin, TX", "Not Started", "", "No", "No", "", "", "", "No", "", "No", "", "Top cybersecurity company", "83", "High"],
    ["", "Atlassian", "Security Engineer Intern", "https://www.atlassian.com/company/careers", "Direct", "Summer 2026", "Seattle/Remote", "Not Started", "", "No", "No", "", "", "", "No", "", "No", "", "$49-75/hr", "85", "High"],
    ["", "CISA", "Cyber/IT Interns", "https://www.cisa.gov/careers", "USAJOBS", "Summer 2026", "Various", "Not Started", "", "No", "No", "", "", "", "No", "", "No", "", "Government, paid", "80", "Medium"],
    ["", "Dex (YC)", "AI/ML Engineer Intern", "https://www.ycombinator.com/companies/joindex", "Y Combinator", "Summer 2026", "San Francisco", "Not Started", "", "No", "No", "", "", "", "No", "", "No", "", "RL, LLMs, AI frameworks", "86", "High"],
    ["", "Mesh (YC)", "AI Engineer Intern", "https://www.ycombinator.com/companies/mesh-2", "Y Combinator", "Summer 2026", "San Francisco", "Not Started", "", "No", "No", "", "", "", "No", "", "No", "", "AI-powered accounting agents", "84", "High"],
    ["", "14 AI (YC)", "Full-Stack Engineering Intern", "https://www.ycombinator.com/companies/14-ai", "Y Combinator", "Summer 2026", "Remote", "Not Started", "", "No", "No", "", "", "", "No", "", "No", "", "AI startup", "82", "High"],
    ["", "Bloom (YC)", "AI/Full-Stack Intern", "https://www.ycombinator.com/companies/bloom-4", "Y Combinator", "Summer 2026", "Zurich", "Not Started", "", "No", "No", "", "", "", "No", "", "No", "", "$6-10K/month", "78", "Medium"],
    ["", "Cohere", "Software Engineering", "https://cohere.com/careers", "Direct", "Summer 2026", "Remote", "Not Started", "", "No", "No", "", "", "", "No", "", "No", "", "AI company", "80", "Medium"],
]

# Color fills
fill_green = PatternFill(start_color="00e5a0", end_color="00e5a0", fill_type="solid")  # Active/Interviewing
fill_yellow = PatternFill(start_color="f5a623", end_color="f5a623", fill_type="solid")  # Applied/Waiting
fill_red = PatternFill(start_color="ff4560", end_color="ff4560", fill_type="solid")  # Rejected
fill_blue = PatternFill(start_color="0077ff", end_color="0077ff", fill_type="solid")  # Offer
fill_gray = PatternFill(start_color="6b7280", end_color="6b7280", fill_type="solid")  # Ghosted

status_colors = {
    "Active": fill_green,
    "Interviewing": fill_green,
    "Applied": fill_yellow,
    "Waiting": fill_yellow,
    "Ready to Submit": fill_yellow,
    "Not Started": fill_yellow,
    "Rejected": fill_red,
    "Offer": fill_blue,
    "Ghosted": fill_gray
}

# Add seed data
for row_num, row_data in enumerate(seed_data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = cell_value
        cell.alignment = Alignment(vertical="center")
        
        # Color code based on status (column 8)
        if col_num == 8 and cell_value in status_colors:
            cell.fill = status_colors[cell_value]
            cell.font = Font(color="FFFFFF" if cell_value in ["Rejected", "Offer", "Ghosted"] else "000000")

# Adjust column widths
column_widths = [12, 20, 25, 35, 15, 12, 15, 15, 18, 12, 15, 18, 25, 12, 12, 12, 15, 15, 30, 12, 10]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Freeze header row
ws.freeze_panes = 'A2'

# Create Summary tab
ws_summary = wb.create_sheet(title="Summary")
ws_summary['A1'] = "INTERNSHIP TRACKER SUMMARY"
ws_summary['A1'].font = Font(bold=True, size=16)
ws_summary.merge_cells('A1:D1')

ws_summary['A3'] = "Metric"
ws_summary['B3'] = "Count"
ws_summary['A3'].font = Font(bold=True)
ws_summary['B3'].font = Font(bold=True)

# Summary formulas
ws_summary['A4'] = "Total Applications"
ws_summary['B4'] = "=COUNTA(Applications!A:A)-1"

ws_summary['A5'] = "Ready to Submit"
ws_summary['B5'] = '=COUNTIF(Applications!H:H,"Ready to Submit")'

ws_summary['A6'] = "Applied"
ws_summary['B6'] = '=COUNTIF(Applications!H:H,"Applied")'

ws_summary['A7'] = "Interviewing"
ws_summary['B7'] = '=COUNTIF(Applications!H:H,"Interviewing")'

ws_summary['A8'] = "Offers"
ws_summary['B8'] = '=COUNTIF(Applications!H:H,"Offer")'

ws_summary['A9'] = "Rejected"
ws_summary['B9'] = '=COUNTIF(Applications!H:H,"Rejected")'

ws_summary['A10'] = "Response Rate"
ws_summary['B10'] = '=IF(B6>0,(B7+B8+B9)/B6*100&"%","N/A")'

ws_summary['A12'] = "High Priority"
ws_summary['B12'] = '=COUNTIF(Applications!U:U,"High")'

ws_summary['A13'] = "Medium Priority"
ws_summary['B13'] = '=COUNTIF(Applications!U:U,"Medium")'

ws_summary['A15'] = "Last Updated"
ws_summary['B15'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Style summary
ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 15

# Save
file_path = "/root/.openclaw/workspace/Anthony_Internship_Tracker_2026.xlsx"
wb.save(file_path)

print(f"✅ Excel tracker created: {file_path}")
print(f"📊 Total rows: {len(seed_data)}")
print(f"📁 Sheets: Applications, Summary")
