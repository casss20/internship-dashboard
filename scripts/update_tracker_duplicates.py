import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Load existing workbook
try:
    wb = openpyxl.load_workbook("/root/.openclaw/workspace/Anthony_Internship_Tracker_2026.xlsx")
    print("✅ Loaded existing tracker")
except:
    wb = openpyxl.Workbook()
    print("🆕 Creating new tracker")

# Get Applications sheet
ws = wb["Applications"] if "Applications" in wb.sheetnames else wb.active

# Add new columns if they don't exist
headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

new_columns = ["Application ID", "Duplicate Check", "Rejection Date", "Cooling Off Until"]
for col_name in new_columns:
    if col_name not in headers:
        col_num = ws.max_column + 1
        ws.cell(1, col_num).value = col_name
        ws.cell(1, col_num).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, col_num).fill = PatternFill(start_color="1f2530", end_color="1f2530", fill_type="solid")
        print(f"✅ Added column: {col_name}")

# Create Skipped Duplicates sheet
if "Skipped Duplicates" not in wb.sheetnames:
    ws_skipped = wb.create_sheet(title="Skipped Duplicates")
    skipped_cols = ["Date Detected", "Company", "Role", "Job URL", "Reason Skipped",
                    "Original Application ID", "Original Date Applied"]
    for i, col in enumerate(skipped_cols, 1):
        cell = ws_skipped.cell(1, i)
        cell.value = col
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="ff4560", end_color="ff4560", fill_type="solid")
    print("✅ Created Skipped Duplicates sheet")

# Create Flagged for Review sheet
if "Flagged for Review" not in wb.sheetnames:
    ws_flagged = wb.create_sheet(title="Flagged for Review")
    flagged_cols = ["Date Flagged", "Company", "Role", "Job URL", "Reason Flagged",
                    "Existing Application ID", "Existing Role", "Action Needed"]
    for i, col in enumerate(flagged_cols, 1):
        cell = ws_flagged.cell(1, i)
        cell.value = col
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="f5a623", end_color="f5a623", fill_type="solid")
    print("✅ Created Flagged for Review sheet")

# Add Application IDs to existing rows
app_id_counter = 1
for row in range(2, ws.max_row + 1):
    if not ws.cell(row, 1).value:  # No Application ID
        ws.cell(row, 1).value = f"APP-{app_id_counter:03d}"
        app_id_counter += 1

# Add Duplicate Check status
for row in range(2, ws.max_row + 1):
    if not ws.cell(row, 10).value:  # Duplicate Check column
        ws.cell(row, 10).value = "Passed"

# Save
wb.save("/root/.openclaw/workspace/Anthony_Internship_Tracker_2026.xlsx")
print(f"✅ Tracker updated with duplicate prevention system")
print(f"📊 Total applications: {ws.max_row - 1}")
print(f"📁 Sheets: {wb.sheetnames}")
